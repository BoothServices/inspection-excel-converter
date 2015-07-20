import openpyxl
from openpyxl.cell import get_column_letter
from sys import argv
#Script, inputFile, outputFile = argv
outputfile = inputfile, ".txt"
def main():
    convert(inputFile)

def extract_areas(ws):
    all_rows = ws.iter_rows("a2:a100")
    areas = [row[0].value for row in all_rows if row[0].value]
    return areas

def extract_issues(ws):
    issues = [cell.value for cell in next(ws.rows)[1:] if cell.value]
    return issues

def grab_column(ws, column_number, areas):
    letter = get_column_letter(column_number)
    cell_range = '{0}2:{0}{1}'.format(letter, len(areas)) # generate B2:B33 or so
    notes = [row[0].value for row in ws.iter_rows(cell_range)]
    return notes

def convert(inputFile):
    wb = openpyxl.load_workbook(filename=inputFile, read_only=True)
    sheet = wb['Sheet1']
    ws = wb.active

    areas = extract_areas(ws)
    issues = extract_issues(ws)
    #print areas
    #print issues

    for idx, issue in enumerate(issues, 2):
        area_notes = grab_column(ws, idx, areas)
        for area, note in zip(areas, area_notes):
            if note:
                print("- {} {} {}".format(issue, area, note))
                with open(outputFile, 'a') as f:
                    f.write("- {} {} {} \n".format(issue, area, note))



if __name__ == "__main__":
    main()
